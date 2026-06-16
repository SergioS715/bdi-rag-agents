"""
Extrae las respuestas y justificaciones de los 3 prompts con mayor
diferencia BASE vs BDI+RAG, uno por rol.
"""
import json
import openpyxl

BASE_DIR = "."

with open(f"{BASE_DIR}/conf/mapeo_condiciones.json", encoding="utf-8") as f:
    mapeo = json.load(f)

wb = openpyxl.load_workbook(f"{BASE_DIR}/plantilla_anotacion.xlsx")

# Prompts con mayor diferencia por rol (calculados previamente)
TARGETS = {
    "victima":    38,
    "victimario": 15,
    "tercero":    34,
}

JUECES_SHEETS = {
    "GPT-5":        "Resultados_GPT5",
    "Gemini Flash": "Resultados_Gemini3Flash",
    "Qwen 36+":     "Resultados_Qwen36Plus",
}

LETRAS = ["A", "B", "C", "D"]
COND_NOMBRES = {
    "baseline": "BASE",
    "rag_only": "RAG",
    "bdi_only": "BDI",
    "bdi_rag":  "BDI+RAG",
}

def get_sheet_row(prompt_num):
    return prompt_num + 1  # fila 1 = cabecera, fila 2 = prompt 1

# Hoja con las respuestas reales
ws_map = wb["Mapeo_Referencia"]

def get_respuesta_letra(prompt_num, letra):
    """Busca la respuesta en la hoja Mapeo_Referencia."""
    col_map = {"A": 3, "B": 4, "C": 5, "D": 6}
    row = prompt_num + 2  # fila 1=cabecera, fila 2=sub-cabecera, fila 3=prompt 1
    val = ws_map.cell(row, col_map[letra]).value
    return val or ""

# -- Intentar leer respuestas desde la hoja de referencia
print("=== ESTRUCTURA HOJA MAPEO_REFERENCIA ===")
for r in range(1, 6):
    row_vals = [ws_map.cell(r, c).value for c in range(1, 7)]
    print(f"Fila {r}: {row_vals}")

print("\n")

for rol, p_num in TARGETS.items():
    clave = f"pregunta_{p_num}"
    info  = mapeo[clave]
    letra_to_cond = info["mapeo"]  # {"A": "baseline", "B": "rag_only", ...}
    cond_to_letra = {v: k for k, v in letra_to_cond.items()}

    print(f"\n{'='*65}")
    print(f"ROL: {rol.upper()}  |  Prompt #{p_num}")
    print(f"Pregunta: {info['pregunta_texto']}")
    print(f"Mapeo letras -> condicion: {letra_to_cond}")
    print()

    # Scores por condicion (promedio 3 jueces)
    scores = {}
    for cond_key in ["baseline", "rag_only", "bdi_only", "bdi_rag"]:
        letra = cond_to_letra[cond_key]
        col_map = {"A": 4, "B": 5, "C": 6, "D": 7}  # D=Punt_A, E=Punt_B...
        row = get_sheet_row(p_num)
        vals = []
        for sheet_name in JUECES_SHEETS.values():
            ws = wb[sheet_name]
            v = ws.cell(row, col_map[letra]).value
            if v is not None:
                vals.append(int(v))
        avg = round(sum(vals)/len(vals), 2) if vals else None
        scores[cond_key] = {"letra": letra, "avg": avg, "vals": vals}
        nombre = COND_NOMBRES[cond_key]
        print(f"  {nombre:8s} (letra {letra}): {avg:.2f}  {vals}")

    print()

    # Justificaciones de GPT-5 (el mas explicativo)
    ws_gpt = wb["Resultados_GPT5"]
    row = get_sheet_row(p_num)
    just = ws_gpt.cell(row, 8).value  # columna H = Justificacion
    print(f"  Justificacion GPT-5:\n  {just}")

    # Respuestas reales por letra
    print()
    print("  Respuestas en Mapeo_Referencia:")
    for letra in LETRAS:
        resp = get_respuesta_letra(p_num, letra)
        cond_key = letra_to_cond.get(letra, "?")
        nombre = COND_NOMBRES.get(cond_key, cond_key)
        print(f"  [{letra} = {nombre}]: {str(resp)[:300]}")
