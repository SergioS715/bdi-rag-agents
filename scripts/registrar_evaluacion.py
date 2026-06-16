#!/usr/bin/env python3
"""
Script para registrar resultados de evaluación LLM en el Excel
Uso: python registrar_evaluacion.py
"""

import openpyxl
from pathlib import Path
import sys

# Configuración
EXCEL_FILE = "plantilla_anotacion.xlsx"
MODELOS = ["ChatGPT", "Gemini", "DeepSeek"]
LETRAS = ["A", "B", "C", "D"]

def validar_ranking(ranking_str):
    """Valida que el ranking sea válido (ej: A > B > C > D)"""
    letras_en_ranking = [x.strip() for x in ranking_str.upper().split(">")]

    # Debe tener 4 letras
    if len(letras_en_ranking) != 4:
        print(f"Error: ranking debe tener 4 letras separadas por >")
        return None

    # Todas deben ser A, B, C, D
    if set(letras_en_ranking) != set(LETRAS):
        print(f"Error: ranking debe contener A, B, C, D (cada una una sola vez)")
        return None

    return " > ".join(letras_en_ranking)

def validar_puntuacion(punt_str):
    """Valida que sea un número entre 1-10"""
    try:
        punt = int(punt_str)
        if 1 <= punt <= 10:
            return punt
        else:
            print("Error: puntuación debe estar entre 1 y 10")
            return None
    except ValueError:
        print("Error: ingresa un número válido")
        return None

def registrar_resultado():
    """Flujo interactivo para registrar un resultado"""
    print("\n" + "="*70)
    print("REGISTRADOR DE EVALUACIONES")
    print("="*70)

    # 1. Seleccionar prompt
    while True:
        try:
            prompt_num = int(input("\nNumero del prompt (1-39): "))
            if 1 <= prompt_num <= 39:
                break
            else:
                print("Error: debe estar entre 1 y 39")
        except ValueError:
            print("Error: ingresa un número válido")

    # 2. Seleccionar modelo
    print("\nModelos disponibles:")
    for i, modelo in enumerate(MODELOS, 1):
        print(f"  {i}. {modelo}")

    while True:
        try:
            modelo_idx = int(input("Elige modelo (1-3): ")) - 1
            if 0 <= modelo_idx < len(MODELOS):
                modelo = MODELOS[modelo_idx]
                break
            else:
                print("Error: elige 1, 2 o 3")
        except ValueError:
            print("Error: ingresa un número válido")

    # 3. Ranking
    while True:
        ranking_input = input("\nRanking (ej: A > C > B > D): ").strip()
        ranking = validar_ranking(ranking_input)
        if ranking:
            break

    # 4. Puntuaciones
    puntuaciones = {}
    print("\nPuntuaciones (1-10 para cada letra):")
    for letra in LETRAS:
        while True:
            punt_input = input(f"  Puntuacion {letra}: ")
            punt = validar_puntuacion(punt_input)
            if punt:
                puntuaciones[letra] = punt
                break

    # 5. Justificación (opcional)
    justificacion = input("\nJustificacion (opcional, presiona Enter para saltar): ").strip()

    # 6. Mostrar resumen
    print("\n" + "-"*70)
    print("RESUMEN:")
    print(f"  Prompt: {prompt_num}")
    print(f"  Modelo: {modelo}")
    print(f"  Ranking: {ranking}")
    print(f"  Puntuaciones: A={puntuaciones['A']}, B={puntuaciones['B']}, C={puntuaciones['C']}, D={puntuaciones['D']}")
    if justificacion:
        print(f"  Justificación: {justificacion}")
    print("-"*70)

    # 7. Confirmar
    while True:
        confirmacion = input("\nEs correcto? (s/n): ").lower()
        if confirmacion in ['s', 'si', 'yes']:
            break
        elif confirmacion in ['n', 'no']:
            print("Cancelado. Vuelve a intentar.")
            return False

    # 8. Escribir en Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)

        # Encontrar la hoja correcta
        sheet_name = f"Resultados_{modelo}"
        if sheet_name not in wb.sheetnames:
            print(f"Error: no existe la hoja {sheet_name}")
            return False

        ws = wb[sheet_name]

        # La fila correcta es: header en fila 1, Prompt 1 en fila 2, etc.
        row = prompt_num + 1

        # Escribir datos
        ws[f'C{row}'] = ranking  # Columna C: Ranking
        ws[f'D{row}'] = puntuaciones['A']  # Columna D: Punt_A
        ws[f'E{row}'] = puntuaciones['B']  # Columna E: Punt_B
        ws[f'F{row}'] = puntuaciones['C']  # Columna F: Punt_C
        ws[f'G{row}'] = puntuaciones['D']  # Columna G: Punt_D
        if justificacion:
            ws[f'H{row}'] = justificacion  # Columna H: Justificación

        # Guardar
        wb.save(EXCEL_FILE)
        print(f"\nGuardado en {sheet_name}, fila {row}")
        return True

    except Exception as e:
        print(f"Error al guardar en Excel: {e}")
        return False

def main():
    # Verificar que existe el archivo
    if not Path(EXCEL_FILE).exists():
        print(f"Error: no existe {EXCEL_FILE}")
        sys.exit(1)

    print("\nBienvenido al registrador de evaluaciones")
    print("(Presiona Ctrl+C para salir en cualquier momento)\n")

    while True:
        try:
            if registrar_resultado():
                print("\n✓ Resultado registrado correctamente\n")

            # Preguntar si continuar
            while True:
                continuar = input("Registrar otro resultado? (s/n): ").lower()
                if continuar in ['s', 'si', 'yes']:
                    break
                elif continuar in ['n', 'no']:
                    print("\nHasta luego!")
                    return
        except KeyboardInterrupt:
            print("\n\nInterrumpido por el usuario.")
            return

if __name__ == "__main__":
    main()
